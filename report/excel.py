from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, Alignment, NamedStyle, colors
import openpyxl.styles as sty
import time

from openpyxl.styles.numbers import FORMAT_DATE_YYYYMMDD2


class Excel:

    def __init__(self):
        self.wb_new = Workbook()
        self.ws_new = self.wb_new.active
        self.ws_new.title = "日报"
        self.yellow_lines = []

    def write_to_excel(self, data):
        self.ws_new.append(['项目', '工作类别', 'Bug ID', '简要描述', '优先级', '是否reopen', 'reopen原因', '解决方案',
                            '原因', '责任人', '日期', '备注'])
        i = 0
        for x in data:
            if x[7] == "入库":
                self.yellow_lines.append(i)
            self.ws_new.append(x)
            i += 1
        self.format_file()
        file_name = "report/temp/" + str(time.time()) + ".xlsx"
        self.wb_new.save(file_name)
        return file_name

    def format_file(self):
        self.ws_new.column_dimensions['A'].width = 20
        self.ws_new.column_dimensions['B'].width = 14
        self.ws_new.column_dimensions['C'].width = 14
        self.ws_new.column_dimensions['D'].width = 70
        self.ws_new.column_dimensions['E'].width = 10
        self.ws_new.column_dimensions['F'].width = 14
        self.ws_new.column_dimensions['G'].width = 14
        self.ws_new.column_dimensions['H'].width = 14
        self.ws_new.column_dimensions['I'].width = 14
        self.ws_new.column_dimensions['J'].width = 14
        self.ws_new.column_dimensions['K'].width = 14
        self.ws_new.column_dimensions['L'].width = 25

        left, right, top, bottom = [Side(style='thin', color='000000')] * 4
        title = NamedStyle(name="title")
        title.font = Font(name=u'宋体', size=11, bold=True)
        title.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        title.border = Border(left=left, right=right, top=top, bottom=bottom)
        content = NamedStyle(name="content")
        content.font = Font(name=u'宋体', size=11)
        content.number_format = FORMAT_DATE_YYYYMMDD2
        content.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        content.border = Border(left=left, right=right, top=top, bottom=bottom)
        content_long = NamedStyle(name="content_long")
        content_long.font = Font(name=u'宋体', size=11)
        content_long.border = Border(left=left, right=right, top=top, bottom=bottom)
        content_long.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        for i in range(self.ws_new.max_row):
            self.ws_new.row_dimensions[i + 1].height = 40

        for x in self.ws_new['A:C']:
            for y in x:
                y.style = content

        for x in self.ws_new['E:F']:
            for y in x:
                y.style = content

        for x in self.ws_new['H:K']:
            for y in x:
                y.style = content

        for x in self.ws_new[1]:
            x.style = title

        for x in self.ws_new['D'][1:]:
            x.style = content_long

        for x in self.ws_new['L'][1:]:
            x.style = content_long

        for x in self.ws_new['G'][1:]:
            x.style = content_long

        for i in self.yellow_lines:
            for x in self.ws_new[i + 2]:
                x.fill = sty.PatternFill(fill_type='solid', fgColor=colors.YELLOW)
