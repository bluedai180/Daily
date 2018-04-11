from django.http import StreamingHttpResponse
from openpyxl import Workbook
from openpyxl.descriptors import Alias
from openpyxl.styles import Border, Side, Font, Alignment, NamedStyle, colors
from django.utils import timezone
import time

from openpyxl.styles import PatternFill
from openpyxl.styles.numbers import FORMAT_DATE_YYYYMMDD2


class Excel:
    def __init__(self):
        self.wb_new = Workbook()
        self.ws_new = self.wb_new.active
        self.yellow_lines = []

    def write_to_excel(self, data, mark=False):
        self.ws_new.append(['项目', '工作类别', 'Bug ID', '简要描述', '优先级', '是否reopen', 'reopen原因', '解决方案',
                            '原因', '责任人', '日期', '备注'])
        self.yellow_lines.clear()
        i = 0
        for x in data:
            if x[7] == "入库":
                self.yellow_lines.append(i)
            self.ws_new.append(x)
            i += 1
        self.format_file()
        if mark:
            self.mark_lines()
        file_name = "report/temp/" + str(time.time()) + ".xlsx"
        self.wb_new.save(file_name)
        return file_name

    def write_weekly_to_excel(self, data, team):
        """写入周报数据，导出excel表格"""
        self.ws_new.append(['项目', '本周工作内容', '工时', '下周工作安排', '疑难问题'])
        for x in data:
            self.ws_new.append(x)
        file_name = "report/temp/" + str(time.time()) + ".xlsx"
        self.ws_new.column_dimensions['A'].width = 20
        self.ws_new.column_dimensions['B'].width = 70
        self.ws_new.column_dimensions['C'].width = 14
        self.ws_new.column_dimensions['D'].width = 70
        self.ws_new.column_dimensions['E'].width = 70
        left, right, top, bottom = [Side(style='thin', color='000000')] * 4
        title = NamedStyle(name="title")
        title.font = Font(name=u'宋体', size=11, bold=True)
        title.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        title.border = Border(left=left, right=right, top=top, bottom=bottom)
        title.fill = PatternFill(fill_type='solid', fgColor="FF0593D3")
        content = NamedStyle(name="content")
        content.font = Font(name=u'宋体', size=11, bold=True)
        content.number_format = FORMAT_DATE_YYYYMMDD2
        content.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        content.border = Border(left=left, right=right, top=top, bottom=bottom)
        content_long = NamedStyle(name="content_long")
        content_long.font = Font(name=u'宋体', size=11)
        content_long.border = Border(left=left, right=right, top=top, bottom=bottom)
        content_long.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        content_long.shrink_to_fit = Alias('shrinkToFit')
        for x in self.ws_new['A']:
            x.style = content
        for x in self.ws_new['B']:
            x.style = content_long
        for x in self.ws_new['C']:
            x.style = content
        for x in self.ws_new['D']:
            x.style = content_long
        for x in self.ws_new['E']:
            x.style = content_long
        for x in self.ws_new[1]:
            x.style = title

        self.ws_new.row_dimensions[1].height = 20
        for i in range(self.ws_new.max_row):
            self.ws_new.row_dimensions[i + 2].height = 150
        self.wb_new.save(file_name)

        def file_iterator(file_name):
            with open(file_name, 'rb') as f:
                while True:
                    c = f.read()
                    if c:
                        yield c
                    else:
                        break

        the_file_name = "weekly-%s-%s.xlsx" % (team, timezone.now().date().strftime("%Y-%m-%d"))
        response = StreamingHttpResponse(file_iterator(file_name))
        response['Content-Type'] = 'application/octet-stream'
        response['Content-Disposition'] = 'attachment;filename="{0}"'.format(the_file_name)
        return response

    def write_hour_allocation_to_excel(self, data):
        self.ws_new.page_setup.fitToHeight = 0
        self.ws_new.page_setup.fitToWidth = 50
        for x in data:
            for index, value in enumerate(x):
                if value == 0:
                    x[index] = ""
            self.ws_new.append(x)
        file_name = "report/temp/" + str(time.time()) + ".xlsx"
        self.wb_new.save(file_name)

        def file_iterator(file_name):
            with open(file_name, 'rb') as f:
                while True:
                    c = f.read()
                    if c:
                        yield c
                    else:
                        break

        the_file_name = "hour_allocation-%s.xlsx" % timezone.now().date().strftime("%Y-%m-%d")
        response = StreamingHttpResponse(file_iterator(file_name))
        response['Content-Type'] = 'application/octet-stream'
        response['Content-Disposition'] = 'attachment;filename="{0}"'.format(the_file_name)
        return response

    def format_file(self):
        """将日报excel文件配置样式"""
        self.ws_new.column_dimensions['A'].width = 20
        self.ws_new.column_dimensions['B'].width = 14
        self.ws_new.column_dimensions['C'].width = 14
        self.ws_new.column_dimensions['D'].width = 70
        self.ws_new.column_dimensions['E'].width = 10
        self.ws_new.column_dimensions['F'].width = 14
        self.ws_new.column_dimensions['G'].width = 14
        self.ws_new.column_dimensions['H'].width = 14
        self.ws_new.column_dimensions['I'].width = 14
        self.ws_new.column_dimensions['J'].width = 14
        self.ws_new.column_dimensions['K'].width = 14
        self.ws_new.column_dimensions['L'].width = 25

        left, right, top, bottom = [Side(style='thin', color='000000')] * 4
        title = NamedStyle(name="title")
        title.font = Font(name=u'宋体', size=11, bold=True)
        title.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        title.border = Border(left=left, right=right, top=top, bottom=bottom)
        content = NamedStyle(name="content")
        content.font = Font(name=u'宋体', size=11)
        content.number_format = FORMAT_DATE_YYYYMMDD2
        content.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        content.border = Border(left=left, right=right, top=top, bottom=bottom)
        content_long = NamedStyle(name="content_long")
        content_long.font = Font(name=u'宋体', size=11)
        content_long.border = Border(left=left, right=right, top=top, bottom=bottom)
        content_long.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        for i in range(self.ws_new.max_row):
            self.ws_new.row_dimensions[i + 1].height = 40

        for x in self.ws_new['A:C']:
            for y in x:
                y.style = content

        for x in self.ws_new['E:F']:
            for y in x:
                y.style = content

        for x in self.ws_new['H:K']:
            for y in x:
                y.style = content

        for x in self.ws_new[1]:
            x.style = title

        for x in self.ws_new['D'][1:]:
            x.style = content_long

        for x in self.ws_new['L'][1:]:
            x.style = content_long

        for x in self.ws_new['G'][1:]:
            x.style = content_long

    def mark_lines(self):
        """有入库信息的行标黄"""
        for i in self.yellow_lines:
            for x in self.ws_new[i + 2]:
                x.fill = PatternFill(fill_type='solid', fgColor=colors.YELLOW)
